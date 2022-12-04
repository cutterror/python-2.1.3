from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import string
from statistic import Statistic
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader


class Report:
    title_font = Font(name='Calibri', size=11, bold=True)

    border = Border(left=Side(border_style="thin", color='FF000000'),
                    right=Side(border_style="thin", color='FF000000'),
                    top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))

    def __init__(self, statistic: Statistic):
        self.__statistic = statistic
        self.__book = Workbook()
        self.__year_list = self.__book.active
        self.__year_list.title = "Статистика по годам"
        self.__city_list = self.__book.create_sheet("Статистика по городам")
        self.fig, self.axs = plt.subplots(nrows=2, ncols=2)
        self.fig.set_figheight(13)
        self.fig.set_figwidth(13)
        plt.rc('axes', titlesize=20)

    def generate_excel(self):
        self.print_columns(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.__statistic.selected_vacancy,
                            'Количество вакансий', 'Количество вакансий - ' + self.__statistic.selected_vacancy],
                           (list(self.__statistic.salary_dynamics.keys()), 'right', False),
                           (list(self.__statistic.salary_dynamics.values()), 'right', False),
                           (list(self.__statistic.selected_salary_dynamics.values()), 'right', False),
                           (list(self.__statistic.num_vacancies_dynamics.values()), 'right', False),
                           (list(self.__statistic.selected_num_vacancies_dynamics.values()), 'right', False))
        self.__book.active = self.__city_list
        self.print_columns(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'],
                           (list(self.__statistic.city_salary_dynamics.keys()), 'left', False),
                           (list(self.__statistic.city_salary_dynamics.values()), 'right', False),
                           (list(), 'right', False),
                           (list(self.__statistic.city_num_vacancies_dynamics.keys()), 'left', False),
                           (list(self.__statistic.city_num_vacancies_dynamics.values()), 'right', True))
        self.__book.active = self.__year_list
        self.__book.save("report.xlsx")

    def print_columns(self, titles: list, *args):
        self.fill_titles(titles)
        columns = string.ascii_uppercase[:len(titles)]
        for i, arg in enumerate(args):
            self.print_column(columns[i], arg[0], arg[1], arg[2])

    def print_column(self, column_name: str, column_data: list, alignment: str, is_percent: bool):
        ws = self.__book.active
        for i in range(len(column_data)):
            ws[column_name + str(i + 2)] = str(round(column_data[i] * 100, 2)) + '%' if is_percent else column_data[i]
            if ws.column_dimensions[column_name].width < len(str(column_data[i])) + 2:
                ws.column_dimensions[column_name].width = len(str(column_data[i])) + 2
            ws[column_name + str(i + 2)].alignment = Alignment(horizontal=alignment)
            ws[column_name + str(i + 2)].border = self.border

    def fill_titles(self, titles: list):
        columns = string.ascii_uppercase[:len(titles)]
        ws = self.__book.active
        for i, column in enumerate(columns):
            if titles[i] == "":
                ws.column_dimensions[column].width = 2
                continue
            ws[column + '1'] = titles[i]
            ws[column + '1'].font = self.title_font
            ws[column + '1'].border = self.border
            ws.column_dimensions[column].width = len(titles[i]) + 2

    def generate_image(self):
        self.create_two_labels_graph(0, 0, self.__statistic.salary_dynamics.values(),
                                     self.__statistic.selected_salary_dynamics.values(), 'средняя з/п',
                                     'з/п ' + self.__statistic.selected_vacancy,
                                     self.__statistic.salary_dynamics.keys(), 'Уровень зарплат по годам')
        self.create_two_labels_graph(0, 1, self.__statistic.num_vacancies_dynamics.values(),
                                     self.__statistic.selected_num_vacancies_dynamics.values(), 'Количество вакансий',
                                     'Количество вакансий ' + '\n' + self.__statistic.selected_vacancy,
                                     self.__statistic.num_vacancies_dynamics.keys(), 'Количество вакансий по годам')
        self.create_horizontal_graph(1, 0, list(self.__statistic.city_salary_dynamics.values()),
                                     list(self.__statistic.city_salary_dynamics.keys()), 'Уровень зарплат по городам')
        other_num_vacancies = 1 - sum(self.__statistic.city_num_vacancies_dynamics.values())
        self.create_pie(1, 1, [other_num_vacancies] + list(self.__statistic.city_num_vacancies_dynamics.values()),
                        ['Другие'] + list(self.__statistic.city_num_vacancies_dynamics.keys()),
                        'Доля вакансий по городам')
        plt.savefig('graph.png', dpi=250)

    def create_two_labels_graph(self, x, y, first_labels, second_labels, first_labels_name,
                                second_labels_name, ticks, title):
        slots = np.arange(len(ticks))
        width = 0.35
        self.axs[x, y].bar(slots - width / 2, first_labels, width, label=first_labels_name)
        self.axs[x, y].bar(slots + width / 2, second_labels, width, label=second_labels_name)
        self.axs[x, y].set_title(title)
        self.axs[x, y].set_xticks(slots, ticks, rotation=90)
        self.axs[x, y].yaxis.grid(visible=True, which='major', color='grey', alpha=.25)
        self.axs[x, y].legend()

    def create_horizontal_graph(self, x, y, labels, ticks, title):
        slots = np.arange(len(ticks))
        labels.reverse()
        ticks.reverse()
        self.axs[x, y].barh(slots, labels, align='center', height=0.5)
        self.axs[x, y].set_title(title)
        self.axs[x, y].set_yticks(slots, ticks)
        self.axs[x, y].xaxis.grid(visible=True, which='major', color='grey', alpha=.25)

    def create_pie(self, x, y, labels, ticks, title):
        self.axs[x, y].pie(labels, labels=ticks)
        self.axs[x, y].set_title(title)

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        years_statistics_trs = self.create_html_years_statistics_trs()
        cities_salary_statistics_trs = self.create_html_cities_statistics_trs(
            self.__statistic.city_salary_dynamics, lambda x: x)
        cities_vacancy_num_statistics_trs = self.create_html_cities_statistics_trs(
            self.__statistic.city_num_vacancies_dynamics, lambda x: str(round(x * 100, 2)) + '%')

        pdf_template = template.render({'selected_vacancy': self.__statistic.selected_vacancy,
                                        'year_statistics_trs': years_statistics_trs,
                                        'city_salary_statistics_trs': cities_salary_statistics_trs,
                                        'city_vacancy_statistics_trs': cities_vacancy_num_statistics_trs})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": True})

    def create_html_years_statistics_trs(self):
        tr_template = ''
        for year in self.__statistic.years:
            average_salary = self.__statistic.salary_dynamics[year]
            selected_average_salary = self.__statistic.selected_salary_dynamics[year]
            vacancies_number = self.__statistic.num_vacancies_dynamics[year]
            selected_vacancies_number = self.__statistic.selected_num_vacancies_dynamics[year]
            tr_template += f'<tr><td class="year">{year}</td><td>{average_salary}</td><td>{selected_average_salary}' + \
                           f'</td><td>{vacancies_number}</td><td>{selected_vacancies_number}</td></tr>'
        return tr_template

    def create_html_cities_statistics_trs(self, city_statistics, format_function):
        tr_template = ''
        for city in city_statistics.keys():
            statistic_value = format_function(city_statistics[city])
            tr_template += f'<tr><td>{city}</td><td>{statistic_value}</td></tr>'
        return tr_template
